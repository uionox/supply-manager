"""Admin area: access control, CRUD, claim moderation, export."""

import io

from openpyxl import load_workbook

from .conftest import ADMIN_PASSWORD


def test_admin_requires_signing_in(client):
    response = client.get("/admin/", follow_redirects=False)
    assert response.status_code == 302
    assert "/admin/login" in response.headers["Location"]


def test_wrong_password_is_refused(client):
    assert b"Incorrect password" in client.post(
        "/admin/login", data={"password": "nope"}).data


def test_sign_out_locks_the_admin_again(admin):
    assert admin.post("/admin/logout", follow_redirects=False).status_code == 302
    response = admin.get("/admin/", follow_redirects=False)
    assert response.status_code == 302 and "login" in response.headers["Location"]


def test_categories_can_be_created_and_reordered(admin, query, scalar):
    admin.post("/admin/categories", data={"name": "Bedding"})
    admin.post("/admin/categories", data={"name": "Food"})
    first = query("SELECT id, name FROM categories ORDER BY sort_order, id")
    assert [row["name"] for row in first] == ["Bedding", "Food"]

    admin.post(f"/admin/categories/{first[0]['id']}/move", data={"direction": "down"})
    after = [r["name"] for r in query("SELECT name FROM categories ORDER BY sort_order, id")]
    assert after == ["Food", "Bedding"]


def test_blank_category_name_is_refused(admin, scalar):
    admin.post("/admin/categories", data={"name": "Bedding"})
    admin.post("/admin/categories", data={"name": "   "})
    assert scalar("SELECT COUNT(*) FROM categories") == 1


def test_item_needs_a_real_category_and_a_positive_quantity(admin, board):
    response = admin.post("/admin/items", data={
        "category_id": board["category_id"], "name": "Bad",
        "quantity_needed": "0", "unit": "pcs"}, follow_redirects=True)
    assert b"at least 1" in response.data

    response = admin.post("/admin/items", data={
        "category_id": 9999, "name": "Orphan",
        "quantity_needed": "5", "unit": "pcs"}, follow_redirects=True)
    assert b"Pick a category" in response.data


def test_target_cannot_drop_below_what_is_claimed(admin, client, board):
    client.post(f"/list/{board['blankets']}", data={"quantity": "6"})
    client.post("/confirm", data={"claimant_name": "Sara", "general_note": ""})

    response = admin.post(f"/admin/items/{board['blankets']}", data={
        "category_id": board["category_id"], "name": "Blankets", "description": "",
        "quantity_needed": "3", "unit": "pcs"}, follow_redirects=True)
    assert b"already claimed" in response.data


def test_admin_can_cancel_a_claim(admin, client, board, scalar):
    client.post(f"/list/{board['blankets']}", data={"quantity": "2"})
    client.post("/confirm", data={"claimant_name": "Sara", "general_note": ""})
    claim_id = scalar("SELECT id FROM claims WHERE claimant_name = 'Sara'")

    response = admin.post(f"/admin/claims/{claim_id}/delete", follow_redirects=True)
    assert b"Claim by Sara cancelled" in response.data
    assert scalar("SELECT COUNT(*) FROM claims") == 0


def test_deleting_an_item_warns_and_takes_its_claims(admin, client, board, scalar):
    client.post(f"/list/{board['blankets']}", data={"quantity": "2"})
    client.post("/confirm", data={"claimant_name": "Sara", "general_note": ""})

    warning = admin.get(f"/admin/items/{board['blankets']}/delete").data
    assert b"Yes, delete" in warning and b"claim(s)" in warning

    admin.post(f"/admin/items/{board['blankets']}/delete", follow_redirects=True)
    assert scalar("SELECT COUNT(*) FROM claims WHERE item_id = ?",
                  (board["blankets"],)) == 0


def test_unknown_item_is_a_404(admin):
    assert admin.get("/admin/items/9999/edit").status_code == 404


def test_dashboard_renders(admin, board):
    response = admin.get("/admin/")
    assert response.status_code == 200 and b"Fulfilled" in response.data


def test_claims_page_shows_notes_and_no_phone(admin, client, board):
    client.post(f"/list/{board['blankets']}", data={"quantity": "2", "note": "size L"})
    client.post("/confirm", data={"claimant_name": "Sara", "general_note": "Friday"})

    page = admin.get(f"/admin/items/{board['blankets']}/claims").data
    assert b"Sara" in page and b"size L" in page and b"Friday" in page
    assert b"Phone" not in page


# ── export ───────────────────────────────────────────────────────────────

def workbook_from(admin):
    response = admin.get("/admin/export.xlsx")
    assert response.status_code == 200
    assert response.data[:2] == b"PK"
    assert "attachment" in response.headers.get("Content-Disposition", "")
    return load_workbook(io.BytesIO(response.data))


def test_export_has_both_sheets_and_split_notes(admin, client, board):
    client.post(f"/list/{board['blankets']}", data={"quantity": "2", "note": "size L"})
    client.post("/confirm", data={"claimant_name": "Sara", "general_note": "Friday"})

    book = workbook_from(admin)
    assert book.sheetnames == ["Claims", "Items"]

    headers = [cell.value for cell in book["Claims"][1]]
    assert "Item note" in headers and "Drop-off note" in headers
    assert "Phone" not in headers

    row = {headers[i]: cell.value for i, cell in enumerate(book["Claims"][2])}
    assert row["Item note"] == "size L"
    assert row["Drop-off note"] == "Friday"
    assert row["Claimant"] == "Sara"


def test_export_has_one_row_per_claim(admin, client, board, scalar):
    client.post(f"/list/{board['blankets']}", data={"quantity": "2"})
    client.post(f"/list/{board['pillows']}", data={"quantity": "1"})
    client.post("/confirm", data={"claimant_name": "Sara", "general_note": ""})

    book = workbook_from(admin)
    assert book["Claims"].max_row == scalar("SELECT COUNT(*) FROM claims") + 1
