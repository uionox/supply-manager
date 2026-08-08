"""Admin area: dashboard, category/item CRUD, claim moderation, Excel export."""

import io
from datetime import date

from flask import (
    Blueprint,
    abort,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .auth import login_required
from .db import get_db, write_transaction
from .timefmt import to_local

bp = Blueprint("admin", __name__, url_prefix="/admin")

XLSX_MIMETYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


@bp.before_request
@login_required
def require_admin():
    """Gate every route in this blueprint behind the admin session."""


# --------------------------------------------------------------------------
# Queries
# --------------------------------------------------------------------------


def _items_with_totals(where="", params=()):
    return get_db().execute(
        f"""
        SELECT i.id, i.category_id, i.name, i.description, i.quantity_needed,
               i.unit, i.sort_order, c.name AS category_name,
               COALESCE(SUM(cl.quantity), 0) AS claimed,
               COUNT(cl.id) AS claim_count
        FROM items i
        JOIN categories c ON c.id = i.category_id
        LEFT JOIN claims cl ON cl.item_id = i.id
        {where}
        GROUP BY i.id
        ORDER BY c.sort_order, c.id, i.sort_order, i.id
        """,
        params,
    ).fetchall()


def _decorate(rows):
    """Add derived remaining/percentage fields to item rows."""
    out = []
    for row in rows:
        needed = row["quantity_needed"]
        claimed = row["claimed"]
        out.append(
            {
                **dict(row),
                "remaining": needed - claimed,
                "percent": round(min(claimed, needed) / needed * 100) if needed else 0,
            }
        )
    return out


def _get_item_or_404(item_id):
    row = get_db().execute(
        """SELECT i.*, c.name AS category_name
           FROM items i JOIN categories c ON c.id = i.category_id
           WHERE i.id = ?""",
        (item_id,),
    ).fetchone()
    if row is None:
        abort(404)
    return row


def _get_category_or_404(cat_id):
    row = get_db().execute("SELECT * FROM categories WHERE id = ?", (cat_id,)).fetchone()
    if row is None:
        abort(404)
    return row


def _reorder(table, row_id, direction, where="1=1", params=()):
    """Move a row one place up or down within its scope.

    Renumbers the whole scope from 0 before swapping, so rows that share a
    sort_order (or were never numbered) still move predictably.
    """
    if direction not in ("up", "down"):
        abort(400)
    with write_transaction() as db:
        # `table` and `where` are module-local literals, never user input.
        ids = [
            r["id"]
            for r in db.execute(
                f"SELECT id FROM {table} WHERE {where} ORDER BY sort_order, id", params
            ).fetchall()
        ]
        if row_id not in ids:
            abort(404)
        pos = ids.index(row_id)
        swap = pos - 1 if direction == "up" else pos + 1
        if 0 <= swap < len(ids):
            ids[pos], ids[swap] = ids[swap], ids[pos]
        for order, ident in enumerate(ids):
            db.execute(f"UPDATE {table} SET sort_order = ? WHERE id = ?", (order, ident))


def _next_sort_order(table, where="1=1", params=()):
    row = get_db().execute(
        f"SELECT COALESCE(MAX(sort_order), -1) + 1 AS next FROM {table} WHERE {where}",
        params,
    ).fetchone()
    return row["next"]


# --------------------------------------------------------------------------
# Dashboard
# --------------------------------------------------------------------------


@bp.get("/")
def dashboard():
    items = _decorate(_items_with_totals())
    total_needed = sum(i["quantity_needed"] for i in items)
    total_claimed = sum(min(i["claimed"], i["quantity_needed"]) for i in items)
    outstanding = sorted(
        (i for i in items if i["remaining"] > 0),
        key=lambda i: (-i["remaining"], i["name"]),
    )
    stats = {
        "total_items": len(items),
        "total_categories": get_db()
        .execute("SELECT COUNT(*) AS n FROM categories")
        .fetchone()["n"],
        "total_claims": get_db()
        .execute("SELECT COUNT(*) AS n FROM claims")
        .fetchone()["n"],
        "fully_claimed": sum(1 for i in items if i["remaining"] <= 0),
        "percent": round(total_claimed / total_needed * 100) if total_needed else 0,
    }
    return render_template(
        "admin/dashboard.html", stats=stats, outstanding=outstanding[:10], items=items
    )


# --------------------------------------------------------------------------
# Categories
# --------------------------------------------------------------------------


@bp.get("/categories")
def categories():
    rows = get_db().execute(
        """SELECT c.id, c.name, c.sort_order, COUNT(i.id) AS item_count
           FROM categories c
           LEFT JOIN items i ON i.category_id = c.id
           GROUP BY c.id
           ORDER BY c.sort_order, c.id"""
    ).fetchall()
    return render_template("admin/categories.html", categories=rows)


@bp.post("/categories")
def create_category():
    name = request.form.get("name", "").strip()
    if not name:
        flash("Category name can't be empty.", "error")
    else:
        get_db().execute(
            "INSERT INTO categories (name, sort_order) VALUES (?, ?)",
            (name, _next_sort_order("categories")),
        )
        flash(f"Category “{name}” added.", "success")
    return redirect(url_for("admin.categories"))


@bp.post("/categories/<int:cat_id>")
def update_category(cat_id):
    _get_category_or_404(cat_id)
    name = request.form.get("name", "").strip()
    if not name:
        flash("Category name can't be empty.", "error")
    else:
        get_db().execute("UPDATE categories SET name = ? WHERE id = ?", (name, cat_id))
        flash("Category renamed.", "success")
    return redirect(url_for("admin.categories"))


@bp.post("/categories/<int:cat_id>/move")
def move_category(cat_id):
    _reorder("categories", cat_id, request.form.get("direction"))
    return redirect(url_for("admin.categories"))


@bp.route("/categories/<int:cat_id>/delete", methods=("GET", "POST"))
def delete_category(cat_id):
    category = _get_category_or_404(cat_id)
    counts = get_db().execute(
        """SELECT COUNT(DISTINCT i.id) AS items, COUNT(cl.id) AS claims
           FROM items i LEFT JOIN claims cl ON cl.item_id = i.id
           WHERE i.category_id = ?""",
        (cat_id,),
    ).fetchone()

    if request.method == "POST":
        get_db().execute("DELETE FROM categories WHERE id = ?", (cat_id,))
        flash(f"Category “{category['name']}” deleted.", "success")
        return redirect(url_for("admin.categories"))

    return render_template(
        "admin/confirm_delete.html",
        title=f"Delete category “{category['name']}”?",
        lines=[
            f"{counts['items']} item(s) in this category will be deleted.",
            f"{counts['claims']} claim(s) recorded against those items will be lost.",
        ],
        action=url_for("admin.delete_category", cat_id=cat_id),
        cancel=url_for("admin.categories"),
    )


# --------------------------------------------------------------------------
# Items
# --------------------------------------------------------------------------


@bp.get("/items")
def items():
    rows = _decorate(_items_with_totals())
    categories_ = get_db().execute(
        "SELECT id, name FROM categories ORDER BY sort_order, id"
    ).fetchall()
    grouped = [
        (cat, [i for i in rows if i["category_id"] == cat["id"]]) for cat in categories_
    ]
    return render_template("admin/items.html", grouped=grouped, has_categories=bool(categories_))


@bp.get("/items/new")
def new_item():
    categories_ = get_db().execute(
        "SELECT id, name FROM categories ORDER BY sort_order, id"
    ).fetchall()
    if not categories_:
        flash("Create a category first.", "error")
        return redirect(url_for("admin.categories"))
    return render_template(
        "admin/item_form.html",
        item=None,
        categories=categories_,
        selected_category=request.args.get("category_id", type=int),
    )


def _read_item_form():
    """Pull item fields off the request. Returns (values, error)."""
    values = {
        "category_id": request.form.get("category_id", type=int),
        "name": request.form.get("name", "").strip(),
        "description": request.form.get("description", "").strip(),
        "unit": request.form.get("unit", "").strip() or "pcs",
        "quantity_needed": request.form.get("quantity_needed", "").strip(),
    }
    if not values["name"]:
        return values, "Item name is required."
    if not values["quantity_needed"].isdigit() or int(values["quantity_needed"]) < 1:
        return values, "Quantity needed must be a whole number of at least 1."
    values["quantity_needed"] = int(values["quantity_needed"])
    exists = get_db().execute(
        "SELECT 1 FROM categories WHERE id = ?", (values["category_id"],)
    ).fetchone()
    if not exists:
        return values, "Pick a category for this item."
    return values, None


@bp.post("/items")
def create_item():
    values, error = _read_item_form()
    if error:
        flash(error, "error")
        return redirect(url_for("admin.new_item"))
    get_db().execute(
        """INSERT INTO items
           (category_id, name, description, quantity_needed, unit, sort_order)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (
            values["category_id"],
            values["name"],
            values["description"],
            values["quantity_needed"],
            values["unit"],
            _next_sort_order("items", "category_id = ?", (values["category_id"],)),
        ),
    )
    flash(f"Item “{values['name']}” added.", "success")
    return redirect(url_for("admin.items"))


@bp.get("/items/<int:item_id>/edit")
def edit_item(item_id):
    item = _get_item_or_404(item_id)
    categories_ = get_db().execute(
        "SELECT id, name FROM categories ORDER BY sort_order, id"
    ).fetchall()
    return render_template(
        "admin/item_form.html",
        item=item,
        categories=categories_,
        selected_category=item["category_id"],
    )


@bp.post("/items/<int:item_id>")
def update_item(item_id):
    item = _get_item_or_404(item_id)
    values, error = _read_item_form()
    if error:
        flash(error, "error")
        return redirect(url_for("admin.edit_item", item_id=item_id))

    claimed = get_db().execute(
        "SELECT COALESCE(SUM(quantity), 0) AS n FROM claims WHERE item_id = ?",
        (item_id,),
    ).fetchone()["n"]
    if values["quantity_needed"] < claimed:
        flash(
            f"{claimed} {item['unit']} are already claimed, so the target can't "
            f"drop below that. Remove some claims first.",
            "error",
        )
        return redirect(url_for("admin.edit_item", item_id=item_id))

    get_db().execute(
        """UPDATE items
           SET category_id = ?, name = ?, description = ?, quantity_needed = ?, unit = ?
           WHERE id = ?""",
        (
            values["category_id"],
            values["name"],
            values["description"],
            values["quantity_needed"],
            values["unit"],
            item_id,
        ),
    )
    flash("Item updated.", "success")
    return redirect(url_for("admin.items"))


@bp.post("/items/<int:item_id>/move")
def move_item(item_id):
    item = _get_item_or_404(item_id)
    _reorder(
        "items",
        item_id,
        request.form.get("direction"),
        "category_id = ?",
        (item["category_id"],),
    )
    return redirect(url_for("admin.items"))


@bp.route("/items/<int:item_id>/delete", methods=("GET", "POST"))
def delete_item(item_id):
    item = _get_item_or_404(item_id)
    if request.method == "POST":
        get_db().execute("DELETE FROM items WHERE id = ?", (item_id,))
        flash(f"Item “{item['name']}” deleted.", "success")
        return redirect(url_for("admin.items"))

    claims = get_db().execute(
        "SELECT COUNT(*) AS n FROM claims WHERE item_id = ?", (item_id,)
    ).fetchone()["n"]
    return render_template(
        "admin/confirm_delete.html",
        title=f"Delete item “{item['name']}”?",
        lines=[f"{claims} claim(s) recorded against this item will be lost."],
        action=url_for("admin.delete_item", item_id=item_id),
        cancel=url_for("admin.items"),
    )


# --------------------------------------------------------------------------
# Claims
# --------------------------------------------------------------------------


@bp.get("/items/<int:item_id>/claims")
def item_claims(item_id):
    item = _decorate(_items_with_totals("WHERE i.id = ?", (item_id,)))
    if not item:
        abort(404)
    claims = get_db().execute(
        """SELECT id, claimant_name, quantity, note, general_note, created_at
           FROM claims WHERE item_id = ? ORDER BY created_at, id""",
        (item_id,),
    ).fetchall()
    return render_template("admin/item_claims.html", item=item[0], claims=claims)


@bp.post("/claims/<int:claim_id>/delete")
def delete_claim(claim_id):
    row = get_db().execute(
        "SELECT item_id, claimant_name FROM claims WHERE id = ?", (claim_id,)
    ).fetchone()
    if row is None:
        abort(404)
    get_db().execute("DELETE FROM claims WHERE id = ?", (claim_id,))
    flash(f"Claim by {row['claimant_name']} cancelled.", "success")
    return redirect(url_for("admin.item_claims", item_id=row["item_id"]))


# --------------------------------------------------------------------------
# Export
# --------------------------------------------------------------------------


def _claimed_at(stored):
    """Camp local time for the spreadsheet.

    Excel cells carry no timezone, so the offset is dropped after converting
    rather than handing over raw UTC that reads hours early.
    """
    moment = to_local(stored, current_app.extensions["display_zone"])
    return moment.replace(tzinfo=None) if moment else stored


def _autofit(sheet, widths):
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_header(sheet, headers, widths):
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    _autofit(sheet, widths)


@bp.get("/export.xlsx")
def export_xlsx():
    items = _decorate(_items_with_totals())
    claims = get_db().execute(
        """SELECT cl.claimant_name, cl.quantity, cl.note, cl.general_note,
                  cl.created_at, i.id AS item_id
           FROM claims cl JOIN items i ON i.id = cl.item_id
           ORDER BY cl.created_at, cl.id"""
    ).fetchall()
    by_id = {i["id"]: i for i in items}

    book = Workbook()

    sheet = book.active
    sheet.title = "Claims"
    _write_header(
        sheet,
        [
            "Category", "Item", "Unit", "Quantity needed", "Quantity remaining",
            "Claimant", "Quantity claimed", "Item note", "Drop-off note",
            "Claimed at",
        ],
        [22, 28, 10, 16, 18, 22, 17, 34, 34, 20],
    )
    for claim in claims:
        item = by_id.get(claim["item_id"], {})
        sheet.append(
            [
                item.get("category_name", ""),
                item.get("name", ""),
                item.get("unit", ""),
                item.get("quantity_needed", ""),
                item.get("remaining", ""),
                claim["claimant_name"],
                claim["quantity"],
                claim["note"] or "",
                claim["general_note"] or "",
                _claimed_at(claim["created_at"]),
            ]
        )

    summary = book.create_sheet("Items")
    _write_header(
        summary,
        ["Category", "Item", "Description", "Unit", "Needed", "Claimed",
         "Remaining", "% fulfilled", "Claims"],
        [22, 28, 46, 10, 10, 10, 12, 13, 9],
    )
    for item in items:
        summary.append(
            [
                item["category_name"], item["name"], item["description"], item["unit"],
                item["quantity_needed"], item["claimed"], item["remaining"],
                item["percent"] / 100, item["claim_count"],
            ]
        )
    for row in summary.iter_rows(min_row=2, min_col=8, max_col=8):
        row[0].number_format = "0%"

    claimed_at_col = 10
    for row in sheet.iter_rows(min_row=2, min_col=claimed_at_col, max_col=claimed_at_col):
        row[0].number_format = "yyyy-mm-dd hh:mm"

    stream = io.BytesIO()
    book.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name=f"supply-tracker-{date.today():%Y-%m-%d}.xlsx",
        mimetype=XLSX_MIMETYPE,
    )
